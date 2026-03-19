#!/usr/bin/env python3
# scripts/ciqual_to_json.py
# Exporte CIQUAL xlsx vers JSON pour le front (sans pandas, openpyxl)
# - Détecte automatiquement la ligne d'entête
# - Exporte code/name + catégories (grp/ssgrp/ssssgrp codes & noms FR)
# - Exporte toutes les colonnes numériques dans nutrients{}
# - Ajoute des alias "canon" (energie_kcal_100g, proteines_g_100g, etc.)

import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional

from openpyxl import load_workbook


DEFAULT_INPUT = "Table Ciqual 2025_FR_2025_11_03.xlsx"
DEFAULT_OUTPUT = "public/ciqual_2025.json"

# ---------------- Candidates columns ----------------
CODE_CANDIDATES = [
    "alim_code", "code", "code_aliment", "ciqual code", "code ciqual", "code aliment"
]
NAME_CANDIDATES = [
    "alim_nom_fr", "nom", "nom_fr", "aliment", "designation", "désignation",
    "denomination", "libelle", "libellé", "nom aliment", "nom de l'aliment"
]

# catégories codes
GRP_CODE_CANDIDATES = ["alim_grp_code", "groupe_code", "grp_code"]
SSGRP_CODE_CANDIDATES = ["alim_ssgrp_code", "sous_groupe_code", "ssgrp_code"]
SSSSGRP_CODE_CANDIDATES = ["alim_ssssgrp_code", "sous_sous_groupe_code", "ssssgrp_code"]

# catégories noms fr
GRP_NAME_CANDIDATES = ["alim_grp_nom_fr", "alim_grp_nom", "groupe_nom_fr", "groupe_nom"]
SSGRP_NAME_CANDIDATES = ["alim_ssgrp_nom_fr", "alim_ssgrp_nom", "sous_groupe_nom_fr", "sous_groupe_nom"]
SSSSGRP_NAME_CANDIDATES = ["alim_ssssgrp_nom_fr", "alim_ssssgrp_nom", "sous_sous_groupe_nom_fr", "sous_sous_groupe_nom"]

# Colonnes meta exactes à exclure des nutriments
EXCLUDE_EXACT = {
    "alim_grp_code",
    "alim_ssgrp_code",
    "alim_ssssgrp_code",
    "alim_grp_nom_fr",
    "alim_ssgrp_nom_fr",
    "alim_ssssgrp_nom_fr",
    "alim_nom_sci",
    "facteur de jones",
    "facteur de jones ",
}

# Exclusion "contient" pour les nutriments (on veut garder les champs catégorie en top-level, pas dans nutrients)
EXCLUDE_CONTAINS = [
    "code", "nom", "designation", "désignation", "libelle", "libellé",
    "groupe", "grp", "ssgrp", "ssssgrp",
]

# ---------------- Normalisation helpers ----------------
def strip_diacritics(s: str) -> str:
    import unicodedata
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )

def norm(s: str) -> str:
    s = strip_diacritics(str(s or "")).lower().strip()
    s = s.replace("’", "'")
    s = re.sub(r"\s+", " ", s)
    return s

def to_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()

def to_float(v: Any) -> Optional[float]:
    """Retourne float si convertible, sinon None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9eE\.\-]+", "", s)
    if s in ("", "-", ".", "-."):
        return None
    try:
        return float(s)
    except:
        return None

def slugify_header(h: str) -> str:
    """
    Convertit une entête CIQUAL en clé stable.
    Ex: "Calcium\n(mg/100 g)" -> "calcium_mg_100g"
    """
    h0 = norm(h)
    h0 = h0.replace("\\n", " ")
    h0 = h0.replace("\n", " ")
    h0 = h0.replace("(", " ").replace(")", " ")
    h0 = h0.replace("/", " ")
    h0 = re.sub(r"[^a-z0-9]+", "_", h0)
    h0 = re.sub(r"_+", "_", h0).strip("_")

    # Harmonisation courante
    h0 = h0.replace("mg_100_g", "mg_100g")
    h0 = h0.replace("g_100_g", "g_100g")
    h0 = h0.replace("ug_100_g", "ug_100g")
    h0 = h0.replace("kj_100_g", "kj_100g")
    h0 = h0.replace("kcal_100_g", "kcal_100g")

    return h0

def find_header_row(ws, max_scan_rows: int = 60) -> Tuple[int, List[str]]:
    """
    Cherche une ligne d'entête : doit contenir un candidat de code + un candidat de name.
    """
    code_c = [norm(x) for x in CODE_CANDIDATES]
    name_c = [norm(x) for x in NAME_CANDIDATES]

    best_row = 1
    best_headers: List[str] = []
    best_score = -1

    for r in range(1, max_scan_rows + 1):
        row = [to_str(c.value) for c in ws[r]]
        row_n = [norm(x) for x in row]

        has_code = any(x in row_n for x in code_c)
        has_name = any(x in row_n for x in name_c)
        score = (4 if has_code else 0) + (4 if has_name else 0)

        for kw in ("energie", "protei", "glucid", "lipid", "eau"):
            if any(kw in x for x in row_n):
                score += 1

        if score > best_score:
            best_score = score
            best_row = r
            best_headers = row

        if score >= 10:
            return r, row

    return best_row, best_headers

def find_col_index(headers: List[str], candidates: List[str]) -> Optional[int]:
    headers_n = [norm(h) for h in headers]
    cand_n = [norm(c) for c in candidates]
    for i, h in enumerate(headers_n):
        if h in cand_n:
            return i
    return None

def should_exclude_header(h: str) -> bool:
    hn = norm(h)
    if not hn:
        return True
    if hn in EXCLUDE_EXACT:
        return True
    for bad in EXCLUDE_CONTAINS:
        # exclude "meta" columns only
        if bad in hn and ("mg" not in hn and "ug" not in hn and "kcal" not in hn and "kj" not in hn and "g" not in hn):
            return True
    return False

# ---------------- Alias mapping (front-friendly keys) ----------------
def build_alias_map(nutrient_cols: List[Tuple[int, str, str]]) -> Dict[str, str]:
    """
    Retourne { alias_key: raw_key } en choisissant la meilleure colonne CIQUAL.
    nutrient_cols: (idx, raw_key, original_header)
    """
    candidates = [raw_key for (_i, raw_key, _h) in nutrient_cols]

    def pick(regex_list: List[str]) -> Optional[str]:
        for rx in regex_list:
            r = re.compile(rx)
            for k in candidates:
                if r.search(k):
                    return k
        return None

    alias: Dict[str, str] = {}

    # Energie kcal
    alias["energie_kcal_100g"] = pick([
        r"energie.*kcal_100g",
        r"energie.*kcal",
    ]) or ""

    # Protéines (souvent "protéines (Nx facteur de Jones)")
    alias["proteines_g_100g"] = pick([
        r"proteines.*g_100g",
        r"proteines_n_x_facteur_de_jones.*g_100g",
    ]) or ""

    alias["lipides_g_100g"] = pick([r"lipides.*g_100g"]) or ""
    alias["glucides_g_100g"] = pick([r"glucides.*g_100g"]) or ""
    alias["fibres_alimentaires_g_100g"] = pick([r"fibres.*g_100g"]) or ""

    alias["calcium_mg_100g"] = pick([r"calcium.*mg_100g"]) or ""
    alias["fer_mg_100g"] = pick([r"\bfer\b.*mg_100g"]) or ""
    alias["sodium_mg_100g"] = pick([r"sodium.*mg_100g"]) or ""
    alias["magnesium_mg_100g"] = pick([r"magnesium.*mg_100g"]) or ""
    alias["potassium_mg_100g"] = pick([r"potassium.*mg_100g"]) or ""

    alias["eau_g_100g"] = pick([r"eau.*g_100g"]) or ""
    alias["sucres_g_100g"] = pick([r"sucres.*g_100g"]) or ""

    # Nettoyage: retirer alias vides
    return {k: v for k, v in alias.items() if v}

# ---------------- Main ----------------
def main():
    inp = Path(sys.argv[1]) if len(sys.argv) >= 2 else Path(DEFAULT_INPUT)
    out = Path(sys.argv[2]) if len(sys.argv) >= 3 else Path(DEFAULT_OUTPUT)
    sheet_name = sys.argv[3] if len(sys.argv) >= 4 else None

    if not inp.exists():
        print(f"[ERROR] Fichier introuvable: {inp}")
        sys.exit(1)

    out.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(inp, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    header_row, headers = find_header_row(ws)
    headers = [to_str(h) for h in headers]

    code_idx = find_col_index(headers, CODE_CANDIDATES)
    name_idx = find_col_index(headers, NAME_CANDIDATES)

    grp_code_idx = find_col_index(headers, GRP_CODE_CANDIDATES)
    ssgrp_code_idx = find_col_index(headers, SSGRP_CODE_CANDIDATES)
    ssssgrp_code_idx = find_col_index(headers, SSSSGRP_CODE_CANDIDATES)

    grp_name_idx = find_col_index(headers, GRP_NAME_CANDIDATES)
    ssgrp_name_idx = find_col_index(headers, SSGRP_NAME_CANDIDATES)
    ssssgrp_name_idx = find_col_index(headers, SSSSGRP_NAME_CANDIDATES)

    if code_idx is None or name_idx is None:
        print("[ERROR] Impossible de trouver les colonnes CODE et/ou NOM.")
        print(f"-> Ligne d'entêtes détectée: {header_row}")
        for i, h in enumerate(headers):
            if h.strip():
                print(f"  [{i:03d}] {repr(h)}")
        sys.exit(2)

    # Nutriments = toutes colonnes convertibles hors meta
    nutrient_cols: List[Tuple[int, str, str]] = []
    for i, h in enumerate(headers):
        if i in (code_idx, name_idx, grp_code_idx, ssgrp_code_idx, ssssgrp_code_idx, grp_name_idx, ssgrp_name_idx, ssssgrp_name_idx):
            continue
        if should_exclude_header(h):
            continue
        key = slugify_header(h)
        if not key:
            continue
        nutrient_cols.append((i, key, h))

    alias_map = build_alias_map(nutrient_cols)

    items: List[dict] = []
    start_row = header_row + 1

    for r in range(start_row, ws.max_row + 1):
        row = ws[r]
        code = to_str(row[code_idx].value)
        name = to_str(row[name_idx].value)

        if not code or not name:
            continue

        nutrients: Dict[str, float] = {}
        for i, key, _orig in nutrient_cols:
            v = to_float(row[i].value)
            if v is None:
                continue
            nutrients[key] = round(v, 6)

        # Ajout alias canon (si la colonne source existe et si pas déjà présent)
        for alias_key, raw_key in alias_map.items():
            if alias_key in nutrients:
                continue
            if raw_key in nutrients:
                nutrients[alias_key] = nutrients[raw_key]

        item = {
            "code": code,
            "name": name,
            "alim_grp_code": to_str(row[grp_code_idx].value) if grp_code_idx is not None else "",
            "alim_ssgrp_code": to_str(row[ssgrp_code_idx].value) if ssgrp_code_idx is not None else "",
            "alim_ssssgrp_code": to_str(row[ssssgrp_code_idx].value) if ssssgrp_code_idx is not None else "",
            "alim_grp_nom_fr": to_str(row[grp_name_idx].value) if grp_name_idx is not None else "",
            "alim_ssgrp_nom_fr": to_str(row[ssgrp_name_idx].value) if ssgrp_name_idx is not None else "",
            "alim_ssssgrp_nom_fr": to_str(row[ssssgrp_name_idx].value) if ssssgrp_name_idx is not None else "",
            "nutrients": nutrients,
        }
        items.append(item)

    out.write_text(json.dumps(items, ensure_ascii=False), encoding="utf-8")

    print(f"[OK] {len(items)} aliments exportés -> {out}")
    print(f"[INFO] Entêtes: ligne {header_row} | code_idx={code_idx} | name_idx={name_idx}")
    print(f"[INFO] Catégories idx: grp_code={grp_code_idx} ssgrp_code={ssgrp_code_idx} ssssgrp_code={ssssgrp_code_idx}")
    print(f"[INFO] Catégories idx: grp_nom={grp_name_idx} ssgrp_nom={ssgrp_name_idx} ssssgrp_nom={ssssgrp_name_idx}")
    print(f"[INFO] Nutriments exportés (colonnes): {len(nutrient_cols)}")
    if alias_map:
        print("[INFO] Aliases ajoutés:")
        for k, v in alias_map.items():
            print(f"  - {k} <= {v}")

    # mini sanity check
    if items:
        ex = items[0]
        print("[INFO] Exemple item keys:", ", ".join(sorted([k for k in ex.keys() if k != "nutrients"])))
        nk = list(ex["nutrients"].keys())[:20]
        print("[INFO] Exemple nutrients keys:", ", ".join(nk))


if __name__ == "__main__":
    main()

